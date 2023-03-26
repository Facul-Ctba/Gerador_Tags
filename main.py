import os
import sys
from time import sleep

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from PySide6.QtCore import QEvent, Qt
from PySide6.QtGui import QIntValidator
from PySide6.QtWidgets import (QAbstractItemView, QApplication, QMainWindow,
                               QTableWidgetItem)
from qt_material import apply_stylesheet

from ui_geratags import Ui_MainWindow


class MainWindow(QMainWindow, Ui_MainWindow):
    def __init__(self):
        super().__init__()
        self.setupUi(self)

        self.limpo = None

        MainWindow.resize(self, 1024, 768)
        self.setWindowTitle("Gerador de Tags")

        self.Save.triggered.connect(self.Save_File)
        self.Exit.triggered.connect(self.close)
        self.Limpar.triggered.connect(self.limpar)
        self.Excluir.triggered.connect(self.excluir)

        self.le_tag.returnPressed.connect(self.set_foco)

        onlyint = QIntValidator(self)
        onlyint.setRange(1, 200)
        self.le_quant.setValidator(onlyint)

        self.tw_tabela.setAlternatingRowColors(True)
        self.tw_tabela.setVerticalScrollMode(QAbstractItemView.ScrollPerItem)

        self.alinhamento = Alignment(horizontal='center', vertical='center')

        self.lista_files()

        self.Mostra_Mensagem('Ready')

    def Open_File(self):
        if self.limpo:
            self.limpo = False
            return
        self.nome_arquivo = self.le_arquivo.text()
        self.nome_arquivo = self.nome_arquivo + ".xlsx"
        existe = os.path.isfile(self.nome_arquivo)
        if existe:
            msg1 = f'...Arquivo aberto: {self.nome_arquivo}'
            self.Mostra_Mensagem(msg1)
            self.arquivo = load_workbook(self.nome_arquivo)
            arq_novo = False
        else:
            msg1 = f'...Criado novo arquivo: {self.nome_arquivo}'
            self.Mostra_Mensagem(msg1)
            self.arquivo = Workbook()
            arq_novo = True

        self.setWindowTitle("Gerador de Tags - " + self.nome_arquivo)

        self.planilha = self.arquivo.active

        if self.planilha.title != 'Tags':
            self.planilha.title = 'Tags'

        for reg in self.planilha.iter_rows():
            for cell in reg:
                cell.alignment = self.alinhamento

        self.tw_tabela.setRowCount(self.planilha.max_row)
        msg2 = msg1 + f'  ( Número de Registros: {self.planilha.max_row} )'
        self.Mostra_Mensagem(msg2)

        # !ATUALIZA A TABLE WIDGET
        self.tw_tabela.clearContents()
        if not arq_novo:
            i = 0
            for row in self.planilha.iter_rows():
                for cell in row:
                    tmp = cell.value
                    valor = QTableWidgetItem(tmp)
                    valor.setTextAlignment(Qt.AlignHCenter)
                    self.tw_tabela.setItem(i, 0, valor)
                    i += 1
            sleep(0.2)
            self.tw_tabela.scrollToBottom()

    def esc_arq(self, arq):
        self.arq = arq
        self.item = self.arq.text()
        self.item = self.item.replace('.xlsx', '')
        self.le_arquivo.setText(self.item)
        self.Open_File()

    def Save_File(self):
        fileName = self.le_arquivo.text()
        fileName = fileName + ".xlsx"
        self.arquivo.save(fileName)
        self.Mostra_Mensagem('...Arquivo Salvo')
        self.lista_files()

    def Add_Item(self):
        self.itens = None
        self.itens = [self.le_tag.text().replace(' ', '')]

        try:
            self.quant = int(self.le_quant.text())
        except ValueError:
            self.quant = 0

        if self.quant > 0 and self.itens[0] != '':
            msg3 = f'...Adicionado {self.quant} tags {self.itens[0]}  \
                - Lembre de Salvar o Arquivo'
            self.Mostra_Mensagem(msg3)
            for item in range(0, self.quant):
                self.planilha.append(self.itens)
                self.atualiza_tabela()
            self.le_tag.setText('')
            self.le_quant.setText('')
            self.le_tag.setFocus()
            sleep(0.2)
            self.tw_tabela.scrollToBottom()

    def atualiza_tabela(self):
        num_linhas = self.tw_tabela.rowCount()
        self.tw_tabela.insertRow(num_linhas)
        novo_item = QTableWidgetItem(self.itens[0])
        novo_item.setTextAlignment(Qt.AlignHCenter)
        self.tw_tabela.setItem(num_linhas, 0, novo_item)

    def limpar(self):
        # !LIMPA A PLANILHA
        num_itens = self.planilha.max_row
        self.planilha.delete_rows(idx=1, amount=num_itens)
        self.Save_File()
        # !LIMPA A TABELA QWIDGET
        self.tw_tabela.clearContents()
        for row in range(1, self.tw_tabela.rowCount()+1):
            self.tw_tabela.removeRow(row)
        self.tw_tabela.setRowCount(self.planilha.max_row)

    def excluir(self):
        os.remove(self.nome_arquivo)
        self.limpo = True
        self.Mostra_Mensagem('...Arquivo Excel Excluído!')
        self.tw_tabela.clearContents()
        self.lista_files()

    def set_foco(self):
        self.le_quant.setFocus()

    def Mostra_Mensagem(self, texto):
        self.texto = texto
        self.statusBar.showMessage(self.texto)

    def lista_files(self):
        files = []
        files.clear()
        self.lw_arquivos.clear()
        for x in os.listdir():
            if x.endswith(".xlsx"):
                files.append(x)
        self.lw_arquivos.addItems(files)

    def Excl_Sel(self):
        try:
            self.selecao = self.tw_tabela.selectedItems()
            indice = self.selecao[0].row()+1
            for item in range(len(self.selecao)):
                self.tw_tabela.removeRow(self.selecao[item].row())

            self.planilha.delete_rows(idx=indice, amount=len(self.selecao))
        except IndexError:
            return


if __name__ == "__main__":
    app = QApplication([])

    app.setStyle('Fusion')
    apply_stylesheet(app, theme='dark_blue.xml')

    window = MainWindow()

    def event(self, event):
        if event.type() == QEvent.KeyPress:
            if event.key() in (Qt.Key_Return, Qt.Key_Enter):
                self.le_tag.setFocus()
            return event(event)

    window.show()
    sys.exit(app.exec())
